#!/usr/bin/env python3
"""Generate an Excel workbook summarising all OCCAM ablation results."""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RESULTS_ROOT = Path(__file__).resolve().parent

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
BEST_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
NUM_FMT_2 = "0.00"
NUM_FMT_4 = "0.0000"


EXPERIMENTS = [
    {
        "id": "Origin_Simulation",
        "description": "原始基线复现 (gyy 环境)",
        "mode": "single",
        "min_mask": 0.0005,
        "max_mask": 0.5,
        "metrics_path": RESULTS_ROOT / "origin_simulation" / "metrics.json",
        "splits": ["val", "test"],
    },
    {
        "id": "OCCAM_Multi",
        "description": "OCCAM-M (multi模式, crop=500, thresh=5/4/3)",
        "mode": "multi",
        "min_mask": 0.0005,
        "max_mask": 0.5,
        "metrics_path": RESULTS_ROOT / "occam_multi" / "results" / "metrics.json",
        "splits": ["val"],
    },
    {
        "id": "A0_baseline",
        "description": "消融基线 (本地环境)",
        "mode": "single",
        "min_mask": 0.0005,
        "max_mask": 0.5,
        "metrics_path": RESULTS_ROOT / "ablation_mask_area" / "results" / "A0_baseline" / "metrics.json",
        "splits": ["val"],
    },
    {
        "id": "A1_min0001",
        "description": "min=0.0001 (更小掩码)",
        "mode": "single",
        "min_mask": 0.0001,
        "max_mask": 0.5,
        "metrics_path": RESULTS_ROOT / "ablation_mask_area" / "results" / "A1_min0001" / "metrics.json",
        "splits": ["val"],
    },
    {
        "id": "A2_min001",
        "description": "min=0.001 (过滤噪声小掩码)",
        "mode": "single",
        "min_mask": 0.001,
        "max_mask": 0.5,
        "metrics_path": RESULTS_ROOT / "ablation_mask_area" / "results" / "A2_min001" / "metrics.json",
        "splits": ["val"],
    },
    {
        "id": "A3_min005",
        "description": "min=0.005 (中等以上目标)",
        "mode": "single",
        "min_mask": 0.005,
        "max_mask": 0.5,
        "metrics_path": RESULTS_ROOT / "ablation_mask_area" / "results" / "A3_min005" / "metrics.json",
        "splits": ["val"],
    },
    {
        "id": "A4_min01",
        "description": "min=0.01 (仅大目标)",
        "mode": "single",
        "min_mask": 0.01,
        "max_mask": 0.5,
        "metrics_path": RESULTS_ROOT / "ablation_mask_area" / "results" / "A4_min01" / "metrics.json",
        "splits": ["val"],
    },
    {
        "id": "A5_max025",
        "description": "max=0.25 (排除大面积)",
        "mode": "single",
        "min_mask": 0.0005,
        "max_mask": 0.25,
        "metrics_path": RESULTS_ROOT / "ablation_mask_area" / "results" / "A5_max025" / "metrics.json",
        "splits": ["val"],
    },
    {
        "id": "A6_max010",
        "description": "max=0.10 (严格排除大掩码)",
        "mode": "single",
        "min_mask": 0.0005,
        "max_mask": 0.10,
        "metrics_path": RESULTS_ROOT / "ablation_mask_area" / "results" / "A6_max010" / "metrics.json",
        "splits": ["val"],
    },
    {
        "id": "A7_tight",
        "description": "min=0.001, max=0.25 (紧凑)",
        "mode": "single",
        "min_mask": 0.001,
        "max_mask": 0.25,
        "metrics_path": RESULTS_ROOT / "ablation_mask_area" / "results" / "A7_tight" / "metrics.json",
        "splits": ["val"],
    },
]


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_subheader_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True, size=10)
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def create_summary_sheet(wb):
    ws = wb.active
    ws.title = "总览"

    headers = [
        "实验ID", "描述", "模式", "min_mask_ratio", "max_mask_ratio",
        "Split", "N", "MAE", "RMSE", "NAE", "Avg Time (s)",
        "MAE [1-10]", "MAE [11-50]", "MAE [51-200]", "MAE [201+]",
        "RMSE [1-10]", "RMSE [11-50]", "RMSE [51-200]", "RMSE [201+]",
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    row = 2
    for exp in EXPERIMENTS:
        if not exp["metrics_path"].exists():
            continue
        data = json.loads(exp["metrics_path"].read_text())
        for split in exp["splits"]:
            m = data.get(split)
            if m is None:
                continue
            by_range = m.get("by_gt_range", {})

            vals = [
                exp["id"], exp["description"], exp["mode"],
                exp["min_mask"], exp["max_mask"],
                split, m["n"], m["mae"], m["rmse"],
                m.get("nae"), m.get("avg_time_sec"),
                by_range.get("1-10", {}).get("mae"),
                by_range.get("11-50", {}).get("mae"),
                by_range.get("51-200", {}).get("mae"),
                by_range.get("201+", {}).get("mae"),
                by_range.get("1-10", {}).get("rmse"),
                by_range.get("11-50", {}).get("rmse"),
                by_range.get("51-200", {}).get("rmse"),
                by_range.get("201+", {}).get("rmse"),
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = THIN_BORDER
                cell.alignment = CENTER
                if isinstance(v, float):
                    cell.number_format = NUM_FMT_4 if col in (10,) else NUM_FMT_2
            row += 1

    ws.freeze_panes = "A2"
    auto_width(ws)

    # Highlight best MAE in val split
    val_rows = []
    for r in range(2, row):
        if ws.cell(row=r, column=6).value == "val":
            val_rows.append(r)
    if val_rows:
        mae_vals = [(ws.cell(row=r, column=8).value, r) for r in val_rows
                     if ws.cell(row=r, column=8).value is not None]
        if mae_vals:
            best_row = min(mae_vals, key=lambda x: x[0])[1]
            for col in range(1, len(headers) + 1):
                ws.cell(row=best_row, column=col).fill = BEST_FILL


def create_per_image_sheet(wb, exp_id, json_path, split):
    sheet_name = f"{exp_id}_{split}"[:31]
    ws = wb.create_sheet(title=sheet_name)

    headers = ["Image", "GT", "Pred", "AE", "SE", "NAE", "Time (s)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    data = json.loads(json_path.read_text())
    for row_idx, item in enumerate(data, 2):
        vals = [
            item["name"], item["gt"], item["pred"], item["ae"],
            item.get("se"), item.get("nae"), item.get("elapsed_sec"),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if isinstance(v, float):
                cell.number_format = NUM_FMT_2

    ws.freeze_panes = "A2"
    auto_width(ws)


def main():
    wb = Workbook()
    create_summary_sheet(wb)

    for exp in EXPERIMENTS:
        base_dir = exp["metrics_path"].parent
        for split in exp["splits"]:
            per_image = base_dir / f"per_image_{split}.json"
            if per_image.exists():
                create_per_image_sheet(wb, exp["id"], per_image, split)

    out_path = RESULTS_ROOT / "OCCAM_experiment_results.xlsx"
    wb.save(str(out_path))
    print(f"Excel saved → {out_path}")


if __name__ == "__main__":
    main()
